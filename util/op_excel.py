import time

from openpyxl import load_workbook
from util.util_fun import Util


class OperateExcel:
    def __init__(self,fn=None):
        if not fn:
            self.fn=Util.get_conf()['testData']
        else:
            self.fn=fn
        self.wb=load_workbook(self.fn)  #打开xlsx文件

    #获取活动单元表，得到的是最后打开的那个表
    def get_data(self):
        ws=self.wb.active
        #a_sheet = self.wb.get_sheet_by_name('Sheet1').active
        return ws

    def get_lines(self):
        return len(list(self.get_data().rows))#获取表的行数

    #获取指定单元格的内容
    def get_value(self,row,col):
        return self.get_data().cell(row=row,column=col).value

    #保存指定内容
    def write_result(self,row,col,value):
        self.get_data().cell(row,col).value=value

    #保存文件
    def save_file(self):
        now=time.strftime('%Y-%m-%d %H %M %S')
        self.wb.save(Util.get_conf()['reportpath']+now+'.xlsx')

if __name__ == '__main__':
    op=OperateExcel()
    print(op.get_value(2,3))