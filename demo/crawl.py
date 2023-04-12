import time

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By


class Crawl:
    def test1(self):
        driver = webdriver.Chrome()
        driver.get('https://robot-service.centaurstech.com/test?appkey=shuiwubl-publicscreen-app-pro&appsecret=94bef0390e31a7350c31686b375edbae&name=')
        time.sleep(5)
        self.wb=load_workbook(r"C:\Users\qiwu-PM\Desktop\crawl\shuiwu\shuiwu.xlsx")
        self.ws = self.wb.active
        self.number=len(list(self.ws.rows))
        print (self.number)
        for i in range(1,self.number+1):
            value=self.ws.cell(i,1).value
            print (value)
            driver.find_element(By.CLASS_NAME,"chat-input").click()
            driver.find_element(By.CLASS_NAME,"chat-input").send_keys(value)
            driver.find_element(By.CLASS_NAME,"chat-input").send_keys(Keys.ENTER)
            time.sleep(2)
        li_list=driver.find_elements(By.CLASS_NAME,"qw-message-content")
        for i in li_list:
            now = time.strftime('%Y-%m-%d %H %M %S')
            li=i.text
            for i in range(1,self.number+1):
                self.ws.cell(i,2).value=li
            self.wb.save(now + '.xlsx')














        #driver.quit()

if __name__ == '__main__':
    Crawl().test1()